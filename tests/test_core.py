"""Core unit tests: ingest, rules, partial-acceptance parsing, report."""

import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook, load_workbook

from batchpilot.config import (FieldRule, Profile, ResponseMap,
                               build_custom_profile, infer_rules,
                               profile_from_dict, profile_to_dict)
from batchpilot.ingest import read_rows
from batchpilot.report import write_report
from batchpilot.rules import validate_rows
from batchpilot.sender import parse_batch_response


def make_profile(**kw):
    defaults = dict(
        key="t", name="t", endpoint="http://x/api",
        fields=[
            FieldRule(name="name", required=True),
            FieldRule(name="email", required=True, type="email", unique=True),
            FieldRule(name="quantity", type="integer", min=0),
        ],
        response_map=ResponseMap(results_path="results", status_field="status",
                                 success_values=["accepted"],
                                 message_field="message", index_field="index"),
    )
    defaults.update(kw)
    return Profile(**defaults)


def xlsx_bytes(rows):
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_ingest_xlsx():
    data = xlsx_bytes([["name", "email", "quantity"],
                       ["Asha", "asha@x.com", 5],
                       [None, None, None],           # blank row skipped
                       ["Ravi", "ravi@x.com", "7"]])
    headers, rows = read_rows(data, "t.xlsx")
    assert headers == ["name", "email", "quantity"]
    assert len(rows) == 2
    assert rows[0]["email"] == "asha@x.com"


def test_ingest_csv():
    headers, rows = read_rows(b"name,email\nA,a@x.com\n,,\n", "t.csv")
    assert headers == ["name", "email"]
    assert len(rows) == 1


def test_rules_catch_errors():
    p = make_profile()
    rows = [
        {"name": "Asha", "email": "asha@x.com", "quantity": 5},      # clean
        {"name": None, "email": "bad-email", "quantity": -2},        # 3 errors
        {"name": "Dup", "email": "asha@x.com", "quantity": "x"},     # dup + type
    ]
    issues = validate_rows(p, rows)
    assert issues[0] == []
    codes1 = {i["code"] for i in issues[1]}
    assert codes1 == {"missing", "format", "range"}
    codes2 = {i["code"] for i in issues[2]}
    assert "duplicate" in codes2 and "type" in codes2


BATCH3 = [{"email": "a@x.com"}, {"email": "b@x.com"}, {"email": "c@x.com"}]


def test_partial_acceptance_parsing():
    p = make_profile()
    body = {"results": [
        {"index": 0, "status": "accepted", "message": ""},
        {"index": 1, "status": "rejected", "message": "bad phone"},
        {"index": 2, "status": "accepted", "message": ""},
    ]}
    out = parse_batch_response(p, body, BATCH3, http_ok=True)
    assert [o["status"] for o in out] == ["success", "failed", "success"]
    assert out[1]["message"] == "bad phone"


def test_whole_batch_fallback():
    p = make_profile()
    out = parse_batch_response(p, {"error": "server exploded"}, BATCH3[:2], http_ok=False)
    assert all(o["status"] == "failed" for o in out)
    out_ok = parse_batch_response(p, {"acknowledged": True}, BATCH3[:2], http_ok=True)
    assert all(o["status"] == "success" for o in out_ok)


def test_fieldassist_style_response():
    """FieldAssist ApiResponse: ResponseList holds only ERRORS keyed by ERPId;
    rows absent from it were accepted."""
    p = make_profile(response_map=ResponseMap(
        results_path="ResponseList", status_field="ResponseStatus",
        success_values=["success", "updated", "created"],
        message_field="Message", match_field="ERPId",
        record_field="OutletErpId", missing_means="success"))
    batch = [{"OutletErpId": "O1", "OutletName": "Shop A"},
             {"OutletErpId": "O2", "OutletName": "Shop B"},
             {"OutletErpId": "O3", "OutletName": "Shop C"}]
    body = {"Message": "done", "Response": "PartialSuccess",
            "ResponseList": [{"ERPId": "O2", "Message": "Beat not found",
                              "ResponseStatus": "Failure", "GUID": "g2"}],
            "ResponseStatusCount": {"Updated": 2, "Failed": 1, "Total": 3}}
    out = parse_batch_response(p, body, batch, http_ok=True)
    assert [o["status"] for o in out] == ["success", "failed", "success"]
    assert out[1]["message"] == "Beat not found"


def test_infer_rules_from_headers():
    rules = infer_rules(["Customer Name", "Email ID", "Phone", "Order Qty",
                         "Unit Price", "Delivery Date"])
    types = {r.name: r.type for r in rules}
    assert types == {"Customer Name": "string", "Email ID": "email",
                     "Phone": "phone", "Order Qty": "integer",
                     "Unit Price": "number", "Delivery Date": "date"}


def test_build_custom_profile_and_roundtrip():
    p = build_custom_profile("https://api.example.com/bulk", "post", "items",
                             "5000", "tok123", ["Email ID", "Order Qty"],
                             index_field="recordIndex")
    assert p.batch_size == 2000  # clamped to the FieldAssist-style maximum
    assert p.headers["Authorization"] == "Bearer tok123"
    assert p.method == "POST" and p.records_key == "items"
    # Blank records key = bare JSON array body (base-script semantics)
    p_bare = build_custom_profile("https://api.example.com/bulk", "POST", "",
                                  "100", "", ["A"], auth_type="basic",
                                  auth_user="u", auth_pass="pw", delay="1")
    assert p_bare.records_key == "" and p_bare.delay == 1.0
    p2 = profile_from_dict(profile_to_dict(p))  # as persisted in job payloads
    assert p2.endpoint == p.endpoint
    assert p2.fields[0].type == "email"
    assert p2.response_map.index_field == "recordIndex"


def test_report_written(tmp_path):
    headers = ["name", "email"]
    rows = [{"name": "A", "email": "a@x.com"}, {"name": "B", "email": "bad"}]
    issues = [[], [{"field": "email", "code": "format",
                    "message": "'email' is not a valid email ('bad')",
                    "severity": "error"}]]
    outcomes = [{"status": "success", "message": ""},
                {"status": "failed", "message": "rejected"}]
    path = tmp_path / "r.xlsx"
    write_report(path, headers, rows, issues, outcomes)
    ws = load_workbook(path).active
    assert ws.max_row == 3
    # Columns: 1=Row, 2=name, 3=email, 4=Validation Flags, 5=API Status, 6=API Message
    assert ws.cell(row=3, column=4).value.startswith("[error]")
    assert ws.cell(row=3, column=5).value == "failed"
