"""Read .xlsx / .csv into a list of dict rows. Headers become JSON keys as-is."""

from __future__ import annotations

import csv
import io
import json
import os
import zipfile
from pathlib import Path

from openpyxl import load_workbook

PAYLOAD_EXTENSIONS = (".json", ".zip")


def is_payload_file(filename: str) -> bool:
    return filename.lower().endswith(PAYLOAD_EXTENSIONS)


def read_payload_files(data: bytes, filename: str) -> list[dict]:
    """For the JSON-files workflow (e.g. per-invoice payloads): a .json file
    (single object or array) or a .zip of .json files.
    Returns [{"name", "payload", "error"}] — payload is the parsed JSON."""
    name = filename.lower()
    out: list[dict] = []
    if name.endswith(".json"):
        try:
            obj = json.loads(data.decode("utf-8-sig"))
            if isinstance(obj, list):
                for i, item in enumerate(obj):
                    out.append({"name": f"{filename} #{i + 1}", "payload": item,
                                "error": None})
            else:
                out.append({"name": filename, "payload": obj, "error": None})
        except ValueError as e:
            out.append({"name": filename, "payload": None,
                        "error": f"invalid JSON: {e}"})
        return out
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        for member in sorted(z.namelist()):
            base = os.path.basename(member)
            if (not member.lower().endswith(".json") or member.startswith("__MACOSX")
                    or base.startswith(".")):
                continue
            try:
                out.append({"name": base,
                            "payload": json.loads(z.read(member).decode("utf-8-sig")),
                            "error": None})
            except ValueError as e:
                out.append({"name": base, "payload": None,
                            "error": f"invalid JSON: {e}"})
    return out


def _clean(value):
    if isinstance(value, str):
        value = value.strip()
        return value if value != "" else None
    return value


def read_rows(path_or_bytes, filename: str) -> tuple[list[str], list[dict]]:
    """Returns (headers, rows). Accepts a Path or raw bytes + filename."""
    data = path_or_bytes.read_bytes() if isinstance(path_or_bytes, Path) else path_or_bytes
    name = filename.lower()
    if name.endswith((".xlsx", ".xlsm", ".xltx")):
        return _read_xlsx(data)
    if name.endswith((".csv", ".tsv")):
        return _read_csv(data, delimiter="\t" if name.endswith(".tsv") else ",")
    raise ValueError(f"Unsupported file type: {filename} (use .xlsx, .csv or .tsv)")


def _read_xlsx(data: bytes) -> tuple[list[str], list[dict]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(h).strip() for h in header_row if h is not None]
    n = len(headers)
    rows = []
    for r in rows_iter:
        if r is None or all(v is None or str(v).strip() == "" for v in r[:n]):
            continue
        rows.append({headers[i]: _clean(r[i]) if i < len(r) else None for i in range(n)})
    wb.close()
    return headers, rows


def _read_csv(data: bytes, delimiter: str = ",") -> tuple[list[str], list[dict]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    try:
        headers = [h.strip() for h in next(reader)]
    except StopIteration:
        return [], []
    rows = []
    for r in reader:
        if not any(cell.strip() for cell in r):
            continue
        rows.append({headers[i]: _clean(r[i]) if i < len(r) else None for i in range(len(headers))})
    return headers, rows
