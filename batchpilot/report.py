"""Per-row Excel report: original data + validation flags + API outcome."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
AMBER = PatternFill("solid", fgColor="FFEB9C")


def write_report(path: Path, headers: list[str], rows: list[dict],
                 issues: list[list[dict]], outcomes: list[dict] | None) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    extra = ["Validation Flags", "API Status", "API Message"]
    all_headers = ["Row"] + headers + extra
    ws.append(all_headers)
    for c in range(1, len(all_headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    for i, row in enumerate(rows):
        flags = "; ".join(f"[{it['severity']}] {it['message']}" for it in issues[i]) if issues else ""
        out = outcomes[i] if outcomes else {"status": "", "message": ""}
        ws.append([i + 1] + [_cell(row.get(h)) for h in headers]
                  + [flags, out.get("status", ""), out.get("message", "")])

        r = i + 2
        status_cell = ws.cell(row=r, column=len(all_headers) - 1)
        if out.get("status") == "success":
            status_cell.fill = GREEN
        elif out.get("status") == "failed":
            status_cell.fill = RED
        elif out.get("status"):
            status_cell.fill = AMBER
        if flags:
            ws.cell(row=r, column=len(all_headers) - 2).fill = AMBER

    for c, h in enumerate(all_headers, start=1):
        width = max(len(str(h)) + 2, 12)
        if h == "Validation Flags":
            width = 45
        elif h == "API Message":
            width = 35
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _cell(v):
    if v is None:
        return ""
    if isinstance(v, (int, float, str, bool)):
        return v
    return str(v)
